import logging
import os
import openpyxl as xl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog

# This is a comment to test GIT
output_file = 'out.xlsx'

logger = logging.getLogger(__name__)
logging.basicConfig(format='%(asctime)s:%(name)s:%(message)s', level=logging.DEBUG)
logger.debug('Start debug level')

filename_materials = ''
filename_SAPs = ''

def convert_string_to_float(string_value):
    try:
        temporal = string_value.replace('.', '')
        temporal = temporal.replace(',', '.')

        return(float(temporal))
    except:
        return 0


def data_mining(data_excel_path, sap_excel_path):
    global input_labour_edit
    global input_ptc_edit
    # OPEN both excel files, and output excel for writing
    data_excel_wb = xl.load_workbook(data_excel_path)
    sap_excel_wb = xl.load_workbook(sap_excel_path)
    output_excel_wb = xl.Workbook()
    sheet_out = output_excel_wb.active

    # Change active worksheet
    sheet = data_excel_wb['Hoja1']
    sheet_sap = sap_excel_wb['Hoja1']
    # Write all material to the new output excel file
    for row in range(1 + 7, sheet.max_row + 1 + 7):
        sheet_out.cell(row, 1).value = sheet.cell(row, 4).value
        component_id = sheet.cell(row, 4).value
        sheet_out.cell(row, 2).value = sheet.cell(row, 5).value
        sheet_out.cell(row, 3).value = sheet.cell(row, 7).value
        sheet_out.cell(row, 4).value = sheet.cell(row, 8).value
        sheet_out.cell(row, 5).value = sheet.cell(row, 9).value

        # Search for the price in the SAP excel
        found = False
        for row2 in range(1, sheet_sap.max_row + 1):
            if sheet_sap.cell(row2, 1).value == component_id:
                try:
                    # sheet_out.cell(row, 6).value = float(sheet_sap.cell(row2, 7).value) * float(sheet_sap.cell(row2, 3).value) / 100_000
                    value1 = convert_string_to_float(sheet_sap.cell(row2, 7).value)
                    value2 = convert_string_to_float(sheet_sap.cell(row2, 3).value)
                    value_suma = (value1 * value2) / 1_000_000
                    sheet_out.cell(row, 6).value = value_suma
                    found = True
                except:
                    sheet_out.cell(row, 6).value = 0
                break

        if not found:
            sheet_out.cell(row, 1).font = Font(bold=True)
            sheet_out.cell(row, 3).font = Font(bold=True)

            pass

    # sheet_out.cell(sheet.max_row + 1 + 2, 5).value = 'MATERIAL'
    # sheet_out.cell(sheet.max_row + 1 + 2, 6).value = '=SUM(F1:F42)'
    #
    # sheet_out.cell(sheet.max_row + 1 + 3, 5).value = 'LABOUR'  # 1,5€
    # sheet_out.cell(sheet.max_row + 1 + 3, 6).value = convert_string_to_float(input_labour_edit.get())
    #
    # sheet_out.cell(sheet.max_row + 1 + 4, 5).value = 'PTC'  # 12€
    # sheet_out.cell(sheet.max_row + 1 + 4, 6).value = convert_string_to_float(input_ptc_edit.get())
    #
    # sheet_out.cell(sheet.max_row + 1 + 5, 5).value = 'HK'
    # sheet_out.cell(sheet.max_row + 1 + 5, 6).value = '=1.051*F' + str(sheet.max_row + 1 + 2) + '+F' + str(
    #     sheet.max_row + 1 + 3)
    #  1.051*MATERIAL + LABOUR
    # sheet_out.cell(sheet.max_row + 1 + 6, 5).value = 'CM2'
    # sheet_out.cell(sheet.max_row + 1 + 6, 6).value = '=100*(F' + str(sheet.max_row + 1 + 4) + '-F' + str(
    #     sheet.max_row + 1 + 5) + ')/F' + str(sheet.max_row + 1 + 5)
    # # 100 * (PTC - HK) / HK


    sheet_out.cell(2, 5).value = 'MATERIAL'
    sheet_out.cell(2, 5).font = Font(bold=True)
    sheet_out.cell(2, 6).value = '=SUM(F7:F' + str(sheet_out.max_row) + ')'
    sheet_out.cell(2, 6).font = Font(bold=True)

    sheet_out.cell(3, 5).value = 'LABOUR'  # 1,5€
    sheet_out.cell(3, 5).font = Font(bold=True)
    sheet_out.cell(3, 6).value = convert_string_to_float(input_labour_edit.get())
    sheet_out.cell(3, 6).font = Font(bold=True)

    sheet_out.cell(4, 5).value = 'PTC'     # 12€
    sheet_out.cell(4, 5).font = Font(bold=True)
    sheet_out.cell(4, 6).value = convert_string_to_float(input_ptc_edit.get())
    sheet_out.cell(4, 6).font = Font(bold=True)

    sheet_out.cell(5, 5).value = 'HK'
    sheet_out.cell(5, 5).font = Font(bold=True)
    sheet_out.cell(5, 6).value = '=1.051*F2+F3'
    sheet_out.cell(5, 6).font = Font(bold=True)

    sheet_out.cell(6, 5).value = 'CM2'
    sheet_out.cell(6, 5).font = Font(bold=True)
    sheet_out.cell(6, 6).value = '=100*(F4-F5)/F5'
    # cm2_value =
    sheet_out.cell(6, 6).font = Font(bold=True)




    sheet_out.column_dimensions['A'].width = 40
    sheet_out.column_dimensions['B'].width = 40
    sheet_out.column_dimensions['C'].width = 40
    sheet_out.column_dimensions['D'].width = 40
    sheet_out.column_dimensions['E'].width = 10
    sheet_out.column_dimensions['F'].width = 10

    try:
        output_excel_wb.save(output_file)
    except:
        logger.error('Cannot save excel. Probably it is opened')


def center_window(root, width, height):
    screen_width = root.winfo_screenwidth()  # 1920x
    screen_height = root.winfo_screenheight()  # 1080

    x_coordinate = (screen_width / 2) - (width / 2)
    y_coordinate = (screen_height / 2) - (height / 2)

    # Positions the window in the center of the page.
    root.geometry("%dx%d+%d+%d" % (width, height, x_coordinate, y_coordinate))


def click_material():
    global filename_materials
    global filename_SAPs
    global run_button
    global input_excel_material

    filename_materials = filedialog.askopenfilename(initialdir=os.getcwd(), title='Choose your Material Excel File',
                                 filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    path, filename = os.path.split(filename_materials)
    logger.info(filename)
    input_excel_material.config(text=filename)

    logger.info('Material =' + filename_materials)
    logger.info('SAP =' + filename_SAPs)

    if filename_materials != '' and filename_SAPs != '':
        run_button.config(state="normal")
        logger.debug('Button to enabled')




def click_SAP():
    global filename_materials
    global filename_SAPs
    global run_button
    global input_excel_sap

    filename_SAPs = filedialog.askopenfilename(initialdir=os.getcwd(), title='Choose your SAP Excel File',
                                                        filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    path, filename = os.path.split(filename_SAPs)
    input_excel_sap.config(text=filename)

    logger.info('Material =' + filename_materials)
    logger.info('SAP =' + filename_SAPs)

    if filename_materials != '' and filename_SAPs != '':
        run_button.config(state="normal")
        logger.debug('Button to enabled')


def click_run():
    global filename_materials
    global filename_SAPs
    global run_label


    #param1 = os.getcwd() + '\dades\datos 2\BOM_715000500_Amica UI.xlsx'
    # param2 = os.getcwd() + '\dades\datos 2\materials_Amica.xlsx'
    run_label.config(text='out.xlsx created, opening.')
    data_mining(filename_materials, filename_SAPs)
    os.startfile(output_file)


def init_GUI():
    global run_button
    global input_excel_material
    global input_excel_sap
    global run_label
    global input_labour_edit
    global input_ptc_edit
    root = tk.Tk()
    root.title('Cost Control')
    # center_window(root, 1920, 1080)
    root.geometry('+%d+%d' % (550, 250))


    #photo = tk.PhotoImage(file='iconexcel.png')
    input_excel_button1 = tk.Button(root, padx=25, anchor='w', text='Select BOM', command=click_material)   # image = photo,
    input_excel_button1.grid(row=1, column=0)
    input_excel_material = tk.Label(root, text='<excel name>')
    input_excel_material.grid(row=1, column=1)
    # input_excel_button1.config(cnf=[...])
    
    input_excel_button2 = tk.Button(root, padx=11, anchor='w', text='Select SAP prices', command=click_SAP)
    input_excel_button2.grid(row=2, column=0)
    input_excel_sap = tk.Label(root, padx=20, text='<excel name>')    
    input_excel_sap.grid(row=2, column=1)

    input_labour_label = tk.Label(root, anchor='w', text='LABOUR=')
    input_labour_label.grid(row=3, column=0)
    input_labour_edit = tk.Entry(root, justify='right')
    input_labour_edit.insert(tk.END, '0,0')
    input_labour_edit.grid(row=3, column=1)


    input_ptc_label = tk.Label(root, anchor='w', text='PTC=')
    input_ptc_label.grid(row=4, column=0)
    input_ptc_edit = tk.Entry(root, justify='right')
    input_ptc_edit.insert(tk.END, '0,0')
    input_ptc_edit.grid(row=4, column=1)

    run_button = tk.Button(root, padx=30, anchor='w', text='Run...', state=tk.DISABLED, command=click_run)
    run_button.grid(row=6, column=0)
    run_label = tk.Label(root, text='')
    run_label.grid(row=6, column=1)

    root.mainloop()


if __name__ == "__main__":
    init_GUI()




    logger.info('MA ends')
